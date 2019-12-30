# -*- coding: utf-8 -*-
# @Author: Anderson
# @Date:   2019-11-11 17:42:18
# @Last Modified by:   Anderson
# @Last Modified time: 2019-12-30 12:39:47
from openpyxl import Workbook
import requests
from bs4 import BeautifulSoup
from urllib.parse import quote_plus
import pdfplumber
import jieba
import os
import re
import json
from pyecharts import options as opts
from pyecharts.globals import ThemeType
from pyecharts.charts import Bar
from pyecharts.charts import WordCloud
from copy import copy
from collections import Counter
from .stop_words import stop_words


def validate_title(title):
	rstr = r"[\/\\\:\*\?\"\<\>\|\%]"
	new_title = re.sub(rstr, "_", title)
	return new_title


class WebCrawlerBot(object):
	"""docstring for WebCrawlerBot"""

	def __init__(self):
		self.headers = {
			'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/72.0.3626.119 Safari/537.36',
		}
		self.weibo_hot_list = []

	def set_cookie(self, cookie):
		self.headers['cookie'] = cookie

	def get_sum_box_office(self):
		req = requests.post("http://cbooo.cn/API/GetData.ashx", data={'MethodName': 'BoxOffice_GetPcHomeList'})
		return req.json()["Data"]['Table'][0]['dapan']

	def get_rank_movie_data(self, index):
		req = requests.post("http://cbooo.cn/API/GetData.ashx", data={'MethodName': 'BoxOffice_GetPcHomeList'})
		if index < len(req.json()["Data"]['Table1']):
			movie_data = req.json()["Data"]['Table1'][index]
			return [
				movie_data["MovieName"],
				movie_data["boxoffice"],
				movie_data["amount"],
				movie_data["releasedate"],
				movie_data["default_url"],
			]
		else:
			raise Exception("排名数字超出范围")

	def get_weibo_hot(self, index):
		# 列表第一位不是正常内容
		index += 1
		if not self.weibo_hot_list:
			base_url = 'https://s.weibo.com/top/summary'
			req = requests.get(base_url, headers=self.headers)
			soup = BeautifulSoup(req.text, 'lxml')
			today_hot = soup.select('#pl_top_realtimehot tr')[1:]
			self.weibo_hot_list = copy(today_hot)
		item = self.weibo_hot_list[index]
		title = item.select('.td-02 a')[0].text.strip()
		hot_count = int(item.select('.td-02 span')[0].text.strip())
		url = item.select('.td-02 a')[0].get('href')
		if 'javascript' in url:
			url = item.select('.td-02 a')[0].get('href_to')
		url = f'https://s.weibo.com{url}'

		# Get detail info
		req = requests.get(url, headers=self.headers)
		soup = BeautifulSoup(req.text, "lxml")
		author = soup.select(".card-wrap .content .info .name")[0].text.strip()
		content = soup.select(".card-wrap .content .txt")[0].text.strip()
		return [hot_count, title, author, content, url]

	def get_liepin(self, keyword, page):
		url = f'https://www.liepin.com/zhaopin/?key={quote_plus(keyword)}&curPage={page}'
		req = requests.get(url, headers=self.headers)
		soup = BeautifulSoup(req.text, 'lxml')
		result = []
		for item in soup.select('.sojob-item-main'):
			job_name = item.select('h3 a')[0].text.strip()
			job_company = item.select('.company-name')[0].text.strip()
			job_field = item.select('.field-financing')[0].text.strip()
			job_salary = item.select('.condition .text-warning')[0].text.strip()
			if job_salary == '面议':
				annual_salary = -1
			else:
				min_salary = int(job_salary[:job_salary.index('-')])
				max_salary = int(job_salary[job_salary.index('-') + 1:job_salary.index('k')])
				months = int(job_salary[job_salary.index('·') + 1:-1])
				annual_salary = (min_salary + max_salary) / 2 * months * 1000
			job_area = item.select('.condition .area')[0].text.strip()
			job_edu = item.select('.condition .edu')[0].text.strip()
			job_experience = item.select('.condition span')[-1].text.strip()
			result.append([job_name, job_company, job_field, job_salary, annual_salary, job_area, job_edu, job_experience])
		return result

	def get_huaban(self, keyword, page, key='k4rwsxf5'):
		url = f"https://huaban.com/search/?q={quote_plus(keyword)}&type=pins&{key}&page={page-1}&per_page=20&wfl=1"
		req = requests.get(url, headers=self.headers)
		source = str(req.text)
		start_index = source.index('app.page["pins"] = ') + len('app.page["pins"] = ')
		end_index = source[start_index:].index('app.page["page"]')
		results = []
		for img in json.loads(source[start_index:start_index + end_index].strip()[:-1]):
			results.append({
				'url': f"https://hbimg.huabanimg.com/{img['file']['key']}",
				'name': validate_title(f"{img['board']['title']}-{img['pin_id']}.jpg")
			})

		return results

	def download_image(self, url, filename, folder):
		if not os.path.exists(folder):
			os.mkdir(folder)
		req = requests.get(url)
		with open(os.path.join(folder, f'{filename}.jpg'), 'wb') as f:
			f.write(req.content)


class ExcelBot(object):
	"""docstring for ExcelBot"""

	def __init__(self):
		self.workbook = Workbook()
		self.sheet = self.workbook.active

	def add_row(self, row):
		try:
			self.sheet.append(row)
		except Exception as e:
			print(e)

	def get_row(self, row):
		data = []
		if isinstance(row, int):
			for cell in self.sheet[row]:
				data.append(cell.value)

		return data

	def get_col(self, col):
		data = []
		if isinstance(col, str):
			for cell in self.sheet[col]:
				data.append(cell.value)
		elif isinstance(col, int):
			# excel column counts from 1
			col += 1
			for row in self.sheet.iter_rows(min_col=col, max_col=col):
				data.append(row[0].value)

		return data

	def clear(self):
		self.workbook.remove_sheet(self.sheet)
		self.sheet = self.workbook.create_sheet('sheet1')

	def save(self, filename):
		self.filename = filename
		self.workbook.save(filename=f'{filename}.xlsx')


class PDFBot(object):
	"""docstring for ExcelBot"""

	def __init__(self):
		self.pdf = None

	def open(self, filename):
		self.pdf = open(pdfplumber.open(filename))

	def save(self, filename):
		self.filename = filename
		self.workbook.save(filename=f'{filename}.xlsx')


class DataAnalysisBot(object):
	"""docstring for DataAnalysisBot"""
	def __init__(self):
		self.data = []

	def set_data(self, data):
		self.data = copy(data)

	def get_word_frequency(self, data, count=20):
		word_frequency = []
		words = ''
		if isinstance(data, str):
			words = data
		elif isinstance(data, list):
			words = '\n'.join([str(item) for item in data])

		punct = set(u''' #:!),.:;?]}¢'"、。〉》」』】〕〗〞︰︱︳﹐､﹒
		﹔﹕﹖﹗﹚﹜﹞！），．：；？｜｝︴︶︸︺︼︾﹀﹂﹄﹏､～￠
		々‖•·ˇˉ―--′’”([{£¥'"‵〈《「『【〔〖（［｛￡￥〝︵︷︹︻
		︽︿﹁﹃﹙﹛﹝（｛“‘-—_…@~/\\''')

		words_cut = list(filter(lambda x: x not in punct, jieba.lcut(words)))
		words_cut = list(filter(lambda x: x not in stop_words, words_cut))
		word_frequency = Counter(words_cut).most_common(count)

		return copy(word_frequency)

	def generate_word_cloud(self, data):
		wordcloud = WordCloud()
		wordcloud.add("", data, word_size_range=[20, 100])
		wordcloud.render('word_cloud.html')

	def generate_bar(self, x_axis, y_axis):
		bar = Bar(
			init_opts=opts.InitOpts(
				width="1280px",
				height="720px"))
		bar.add_xaxis(x_axis)
		bar.add_yaxis("", y_axis)
		bar.set_global_opts(
			xaxis_opts=opts.AxisOpts(axislabel_opts=opts.LabelOpts(rotate=10)))
		bar.render('bar.html')


pachong = wc_bot = web_crawler_bot = WebCrawlerBot()
biaoge = ec_bot = excel_bot = ExcelBot()
shuju = da_bot = data_analysis_bot = DataAnalysisBot()
pdf_bot = PDFBot()
